"""
Dad Reimbursement — auto-select credit card transactions for monthly expense splitting.

Priority tiers:
  ALWAYS:  Gas, Groceries, Education
  LIMITED: Fast food / restaurants — max 6 per month
  NEVER:   Everything else
"""

import io
import uuid
from datetime import date, timedelta
from decimal import Decimal
from typing import Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from database import get_db
from dependencies import verify_api_key
from models.db_models import Account, Transaction

router = APIRouter(prefix="/reimbursement", tags=["reimbursement"])

# ---------------------------------------------------------------------------
# Tier classification
# ---------------------------------------------------------------------------

# (category, subcategory) → tier
# "always" = always include, "limited" = capped at MAX_LIMITED, "never" = excluded
MAX_LIMITED = 6

def _classify(category: str | None, subcategory: str | None) -> str:
    cat = (category or "").lower()
    sub = (subcategory or "").lower()

    # Always: gas
    if cat == "transport" and sub == "gas":
        return "always"
    # Always: groceries
    if cat == "food & drink" and sub == "groceries":
        return "always"
    # Always: education
    if cat == "education":
        return "always"
    # Limited: restaurants / coffee / fast food
    if cat == "food & drink" and sub in ("restaurants", "coffee"):
        return "limited"
    # Everything else
    return "never"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _month_bounds(month_str: str) -> tuple[date, date]:
    year, mon = int(month_str[:4]), int(month_str[5:7])
    first_day = date(year, mon, 1)
    if mon == 12:
        last_day = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        last_day = date(year, mon + 1, 1) - timedelta(days=1)
    return first_day, last_day


def _current_month() -> str:
    return date.today().strftime("%Y-%m")


# ---------------------------------------------------------------------------
# Schemas
# ---------------------------------------------------------------------------

class ReimbursementTransaction(BaseModel):
    id: str
    date: date
    amount: float
    merchant_name: Optional[str]
    category: Optional[str]
    subcategory: Optional[str]
    account_alias: str
    tier: str          # "always", "limited", "never"
    auto_selected: bool

    class Config:
        from_attributes = True


class ReimbursementResponse(BaseModel):
    month: str
    transactions: list[ReimbursementTransaction]
    auto_selected_total: float
    rent_amount: float
    misc_target: float
    grand_total: float


class ExportRequest(BaseModel):
    transaction_ids: list[str]
    rent_amount: float = 750
    month: str
    verbose: bool = False


class EmailRequest(BaseModel):
    transaction_ids: list[str]
    rent_amount: float = 750
    month: str
    verbose: bool = False


class CategoryTier(BaseModel):
    category: str
    subcategory: str
    tier: str


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------

@router.get("/transactions", response_model=ReimbursementResponse)
def get_reimbursement_transactions(
    month: str = Query(default=None),
    rent_amount: float = Query(default=750),
    misc_target: float = Query(default=500),
    db: Session = Depends(get_db),
    _: str = Depends(verify_api_key),
):
    if month is None:
        month = _current_month()

    first_day, last_day = _month_bounds(month)

    # Get ALL transactions for the month (all accounts)
    rows = (
        db.query(Transaction, Account.alias)
        .join(Account, Transaction.account_id == Account.id)
        .filter(
            Transaction.date >= first_day,
            Transaction.date <= last_day,
            Transaction.is_pending == False,
            Transaction.amount > 0,  # Only debits (money spent)
        )
        .order_by(Transaction.date.desc())
        .all()
    )

    # Classify and build transaction list
    always_txns = []
    limited_txns = []
    never_txns = []

    for txn, alias in rows:
        tier = _classify(txn.category, txn.subcategory)
        item = ReimbursementTransaction(
            id=str(txn.id),
            date=txn.date,
            amount=float(txn.amount),
            merchant_name=txn.merchant_name,
            category=txn.category,
            subcategory=txn.subcategory,
            account_alias=alias,
            tier=tier,
            auto_selected=False,
        )
        if tier == "always":
            always_txns.append(item)
        elif tier == "limited":
            limited_txns.append(item)
        else:
            never_txns.append(item)

    # Auto-select: always include tier-1
    for t in always_txns:
        t.auto_selected = True

    # Sort limited by amount desc, pick up to MAX_LIMITED
    limited_txns.sort(key=lambda t: t.amount, reverse=True)
    always_total = sum(t.amount for t in always_txns)
    remaining_budget = misc_target - always_total

    selected_limited = 0
    for t in limited_txns:
        if selected_limited >= MAX_LIMITED:
            break
        if remaining_budget <= 0:
            break
        t.auto_selected = True
        remaining_budget -= t.amount
        selected_limited += 1

    all_txns = always_txns + limited_txns + never_txns
    auto_total = sum(t.amount for t in all_txns if t.auto_selected)
    grand_total = rent_amount + auto_total

    return ReimbursementResponse(
        month=month,
        transactions=all_txns,
        auto_selected_total=round(auto_total, 2),
        rent_amount=rent_amount,
        misc_target=misc_target,
        grand_total=round(grand_total, 2),
    )


@router.get("/categories")
def get_category_tiers(_: str = Depends(verify_api_key)):
    """Return the category-to-tier mapping."""
    return [
        {"category": "Transport", "subcategory": "Gas", "tier": "always"},
        {"category": "Food & Drink", "subcategory": "Groceries", "tier": "always"},
        {"category": "Education", "subcategory": "*", "tier": "always"},
        {"category": "Food & Drink", "subcategory": "Restaurants", "tier": "limited"},
        {"category": "Food & Drink", "subcategory": "Coffee", "tier": "limited"},
        {"category": "Shopping", "subcategory": "*", "tier": "never"},
        {"category": "Entertainment", "subcategory": "*", "tier": "never"},
        {"category": "Health", "subcategory": "*", "tier": "never"},
        {"category": "Other", "subcategory": "*", "tier": "never"},
    ]


def _build_xlsx(db: Session, transaction_ids: list[str], rent_amount: float, month: str, verbose: bool) -> tuple[io.BytesIO, str]:
    """Build the xlsx and return (buffer, filename)."""
    import openpyxl
    import calendar
    from openpyxl.styles import Font

    txn_ids = [uuid.UUID(tid) for tid in transaction_ids]
    rows = (
        db.query(Transaction, Account.alias)
        .join(Account, Transaction.account_id == Account.id)
        .filter(Transaction.id.in_(txn_ids))
        .order_by(Transaction.date)
        .all()
    )

    # Group by display category (simplified for spreadsheet)
    def _spreadsheet_category(cat: str | None, sub: str | None) -> str:
        cat = cat or "other"
        sub = sub or ""
        if cat == "Transport" and sub == "Gas":
            return "gas"
        if cat == "Food & Drink" and sub == "Groceries":
            return "groceries"
        if cat == "Food & Drink":
            return "fast food"
        if cat == "Education":
            return "edu"
        return cat.lower()

    from collections import OrderedDict
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font

    # Build category → list of transaction data
    categories: dict[str, list[dict]] = OrderedDict()

    # Rent always first
    categories["rent"] = [{"amount": rent_amount, "merchant": "Rent", "date": None, "account": ""}]

    for txn, alias in rows:
        cat_name = _spreadsheet_category(txn.category, txn.subcategory)
        if cat_name not in categories:
            categories[cat_name] = []
        categories[cat_name].append({
            "amount": float(txn.amount),
            "merchant": txn.merchant_name or "Unknown",
            "date": txn.date.strftime("%m/%d") if txn.date else "",
            "account": alias,
        })

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    bold = Font(bold=True)
    currency_fmt = '_("$"* #,##0.00_);_("$"* \\(#,##0.00\\);_("$"* "-"??_);_(@_)'

    if not verbose:
        # --- Simple mode: category headers + amounts only ---
        col_names = list(categories.keys())
        for col_idx, name in enumerate(col_names, start=1):
            ws.cell(row=1, column=col_idx, value=name)

        max_rows = max(len(v) for v in categories.values())
        for col_idx, name in enumerate(col_names, start=1):
            for row_idx, entry in enumerate(categories[name], start=2):
                cell = ws.cell(row=row_idx, column=col_idx, value=entry["amount"])
                cell.number_format = currency_fmt

        total_row = max_rows + 3
        last_data_row = max_rows + 1
        total_col = len(col_names) + 2

        ws.cell(row=total_row - 1, column=total_col, value="total transfer")
        first_col_letter = get_column_letter(1)
        last_col_letter = get_column_letter(len(col_names))
        formula = f"=SUM({first_col_letter}2:{last_col_letter}{last_data_row})"
        total_cell = ws.cell(row=total_row, column=total_col, value=formula)
        total_cell.number_format = '_("$"* #,##0.00_);_("$"* \\(#,##0.00\\);_("$"* "-"??_);_(@_)'

    else:
        # --- Verbose mode: each category gets 3 columns (merchant | date | amount) ---
        col_names = list(categories.keys())
        cols_per_cat = 3  # merchant, date, amount

        # Row 1: category headers (merged across 3 cols visually, written in amount col)
        for cat_idx, name in enumerate(col_names):
            base_col = cat_idx * cols_per_cat + 1
            cell = ws.cell(row=1, column=base_col, value=name)
            cell.font = bold
            # Sub-headers
            ws.cell(row=2, column=base_col, value="merchant").font = Font(italic=True, color="808080")
            ws.cell(row=2, column=base_col + 1, value="date").font = Font(italic=True, color="808080")
            ws.cell(row=2, column=base_col + 2, value="amount").font = Font(italic=True, color="808080")

        # Row 3+: transaction data
        max_rows = max(len(v) for v in categories.values())
        for cat_idx, name in enumerate(col_names):
            base_col = cat_idx * cols_per_cat + 1
            for row_offset, entry in enumerate(categories[name]):
                data_row = row_offset + 3
                ws.cell(row=data_row, column=base_col, value=entry["merchant"])
                ws.cell(row=data_row, column=base_col + 1, value=entry["date"])
                amount_cell = ws.cell(row=data_row, column=base_col + 2, value=entry["amount"])
                amount_cell.number_format = currency_fmt

        # Total row
        last_data_row = max_rows + 2
        total_row = last_data_row + 2
        total_col = len(col_names) * cols_per_cat + 2

        ws.cell(row=total_row - 1, column=total_col, value="total transfer").font = bold

        # Sum all amount columns (every 3rd column starting from col 3)
        sum_parts = []
        for cat_idx in range(len(col_names)):
            amount_col = get_column_letter(cat_idx * cols_per_cat + 3)
            sum_parts.append(f"{amount_col}3:{amount_col}{last_data_row}")
        formula = "=" + "+".join(f"SUM({p})" for p in sum_parts)
        total_cell = ws.cell(row=total_row, column=total_col, value=formula)
        total_cell.number_format = '_("$"* #,##0.00_);_("$"* \\(#,##0.00\\);_("$"* "-"??_);_(@_)'

        # Auto-width for merchant columns
        for cat_idx in range(len(col_names)):
            merchant_col = get_column_letter(cat_idx * cols_per_cat + 1)
            ws.column_dimensions[merchant_col].width = 22
            date_col = get_column_letter(cat_idx * cols_per_cat + 2)
            ws.column_dimensions[date_col].width = 8

    # Save to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    year, mon = int(month[:4]), int(month[5:7])
    month_name = calendar.month_name[mon]
    short_year = str(year)[2:]
    filename = f"{month_name} '{short_year} expenses.xlsx"
    return buf, filename


@router.post("/export")
def export_reimbursement(
    payload: ExportRequest,
    db: Session = Depends(get_db),
    _: str = Depends(verify_api_key),
):
    """Generate xlsx and return as download."""
    buf, filename = _build_xlsx(db, payload.transaction_ids, payload.rent_amount, payload.month, payload.verbose)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.spreadsheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/email")
def email_reimbursement(
    payload: EmailRequest,
    db: Session = Depends(get_db),
    _: str = Depends(verify_api_key),
):
    """Generate xlsx and email it to dad."""
    import calendar
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.base import MIMEBase
    from email import encoders

    buf, filename = _build_xlsx(db, payload.transaction_ids, payload.rent_amount, payload.month, payload.verbose)

    year, mon = int(payload.month[:4]), int(payload.month[5:7])
    month_name = calendar.month_name[mon]

    # Build email
    msg = MIMEMultipart()
    msg["From"] = "noreply@nathanblatter.com"
    msg["To"] = "eric.blatter296@gmail.com"
    msg["Subject"] = filename.replace(".xlsx", "")

    body = f"Hi Dad,\n\nHere are my expenses for {month_name} {year}.\n\nThanks,\nNathan"
    msg.attach(MIMEText(body, "plain"))

    # Attach xlsx
    part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.spreadsheet")
    part.set_payload(buf.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f'attachment; filename="{filename}"')
    msg.attach(part)

    # Send via local postfix
    try:
        with smtplib.SMTP("docker-services-postfix-1", 25, timeout=10) as smtp:
            smtp.sendmail(msg["From"], [msg["To"]], msg.as_string())
    except Exception as exc:
        from fastapi import HTTPException, status
        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail=f"Failed to send email: {exc}")

    return {"status": "sent", "to": msg["To"], "subject": msg["Subject"]}
